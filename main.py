import json
import time
import csv
import pandas as pd
import os
import openpyxl
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

row1 = 2
row149 = 2
row97 = 2
row98 = 2
row100 = 2
row152 = 2
row8 = 2
row99 = 2
rowhistid = 2
rowhistlabel = 2
rowhistpessoa = 2
rowhistdata = 2


# Load user authentication information from an outside file
with open('C:/OneDrives/OneDrive - UniCEUB/WorkSpaces/newTest/auth.json') as f:
    auth = json.load(f)

# Launch a web browser and navigate to the login page
driver = webdriver.Edge()
driver.get('https://eaud-t.cgu.gov.br/oauth2/authorization/adfs?')

# Enter email and click "Next"
email_field = driver.find_element(By.ID, 'emailInput')
email_field.send_keys(auth['username'])
email_field.send_keys(Keys.RETURN)

# Wait for the page to load and enter username and password
wait = WebDriverWait(driver, 10)
username_field = wait.until(EC.presence_of_element_located((By.ID, 'userNameArea')))
password_field = driver.find_element(By.ID, 'passwordInput')
password_field.send_keys(auth['password'])
password_field.send_keys(Keys.RETURN)

# Navigate to the desired page and start scraping
driver.get('https://eaud-t.cgu.gov.br/auth/monitoramento?')

# Wait for ID='carregando' to disappear
wait = WebDriverWait(driver, 50)
wait.until(EC.invisibility_of_element_located((By.ID, 'carregando')))

while True:
    # Wait for the "carregando" ID to disappear
    wait.until(EC.invisibility_of_element_located((By.ID, 'carregando')))

    # Wait for the page to finish loading
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
    WebDriverWait(driver, 30).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')

    # Get all the <td class="sorting_1"> elements
    td_elements = driver.find_elements(By.CSS_SELECTOR, 'td.sorting_1')

    for td_element in td_elements:
        # Get the href from the <a> element inside the <td>
        href = td_element.find_element(By.CSS_SELECTOR, 'a').get_attribute('href')

        # Open the page in a new tab
        driver.execute_script("window.open('" + href + "', '_blank');")

        # Switch to the new tab
        driver.switch_to.window(driver.window_handles[-1])

        # Wait for the "carregando" ID to disappear
        wait.until(EC.invisibility_of_element_located((By.ID, 'carregando')))

        # Wait for the page to finish loading
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
        WebDriverWait(driver, 30).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')

        # Load the workbook
        workbook = load_workbook(filename="monitoramentoCompleto.xlsx")
        worksheet= workbook.active
        worksheet = workbook['Monitoramento']

##############################################################################################
        try:
            # Find the div with id "divExibicao_-1"
            div = driver.find_element(By.ID, "divExibicao_-1")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "campo-texto-quebra")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row1, column=2).value:
                    row1 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row1, column=2, value=text)

                # Increment row number for next write
                row1 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row1, column=2).value:
                row1 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row1, column=2, value='N/A')

            # Increment row number for next write
            row1 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")

##############################################################################################

        try:
            # Find the div with id "divExibicao_149"
            div = driver.find_element(By.ID, "divExibicao_149")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "campo-texto-quebra")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row149, column=3).value:
                    row149 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row149, column=3, value=text)

                # Increment row number for next write
                row149 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row149, column=3).value:
                row149 += 1
               
            # Write the text to the cell
            worksheet.cell(row=row149, column=3, value='N/A')

            # Increment row number for next write
            row149 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")

##############################################################################################

        try:
            # Find the div with id "divExibicao_97"
            div = driver.find_element(By.ID, "divExibicao_97")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "form-control.campo-form")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row97, column=4).value:
                    row97 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row97, column=4, value=text)

                # Increment row number for next write
                row97 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row97, column=4).value:
                row97 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row97, column=4, value='N/A')

            # Increment row number for next write
            row97 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")

##############################################################################################

        try:
            # Find the div with id "divExibicao_98"
            div = driver.find_element(By.ID, "divExibicao_98")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "campo-form.lista-valores-campo")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row98, column=5).value:
                    row98 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row98, column=5, value=text)

                # Increment row number for next write
                row98 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row98, column=5).value:
                row98 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row98, column=5, value='N/A')

            # Increment row number for next write
            row98 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")


##############################################################################################

        try:
            # Find the div with id "divExibicao_152"
            div = driver.find_element(By.ID, "divExibicao_152")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "form-control.campo-form")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row152, column=6).value:
                    row152 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row152, column=6, value=text)

                # Increment row number for next write
                row152 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row152, column=6).value:
                row152 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row152, column=6, value='N/A')

            # Increment row number for next write
            row152 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx") 

##############################################################################################

        try:
            # Find the div with id "divExibicao_100"
            div = driver.find_element(By.ID, "divExibicao_100")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "campo-form.lista-valores-campo")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row100, column=7).value:
                    row100 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row100, column=7, value=text)

                # Increment row number for next write
                row100 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row100, column=7).value:
                row100 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row100, column=7, value='N/A')

            # Increment row number for next write
            row100 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")

##############################################################################################

        try:
            # Find the div with id "divExibicao_99"
            div = driver.find_element(By.ID, "divExibicao_99")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "form-control.campo-form")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row99, column=8).value:
                    row99 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row99, column=8, value=text)

                # Increment row number for next write
                row99 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row99, column=8).value:
                row99 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row99, column=8, value='N/A')

            # Increment row number for next write
            row99 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")

##############################################################################################

        try:
            # Find the div with id "divExibicao_-8"
            div = driver.find_element(By.ID, "divExibicao_-8")

            # Find the pre element with class "campo-texto-quebra" inside the div
            pre = div.find_element(By.CLASS_NAME, "campo-form.lista-valores-campo")
            text = pre.text.strip()

            # Write the text to the worksheet
            if text:
                # Find the next empty cell in the first column
                while worksheet.cell(row=row8, column=9).value:
                    row8 += 1
                
                # Write the text to the cell
                worksheet.cell(row=row8, column=9, value=text)

                # Increment row number for next write
                row8 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")
        except NoSuchElementException:
            # Find the next empty cell in the first column
            while worksheet.cell(row=row8, column=9).value:
                row8 += 1
                
            # Write the text to the cell
            worksheet.cell(row=row8, column=9, value='N/A')

            # Increment row number for next write
            row8 += 1

            # Save the workbook
            workbook.save("monitoramentoCompleto.xlsx")


##############################################################################################

        # Find the "next" button
        exibir_mais = driver.find_element(By.CSS_SELECTOR, "button#btn-exibir-mais.btn.btn-success.btn-labeled")
        exibir_mais.click()

        # Load second WorkSheet
        worksheet = workbook['Historico']

##############################################################################################

        ## Add "Histórico" (ToBe Completed)

        # Wait for the "carregando" ID to disappear
        wait.until(EC.invisibility_of_element_located((By.ID, 'carregando')))

        # Wait for the page to finish loading
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
        WebDriverWait(driver, 30).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')

        # Get all the <td class="sorting_1"> elements
        hists = driver.find_elements(By.CSS_SELECTOR, 'div.interacoes')

        for hist in hists:

            try:    
                id = driver.find_element(By.CSS_SELECTOR, 'span.label.label-id.label-lg')
                idtext = id.text.strip()

                # Write the text to the worksheet
                if idtext:
                    # Find the next empty cell in the first column
                    while worksheet.cell(row=rowhistid, column=1).value:
                        rowhistid += 1
                    
                    # Write the text to the cell
                    worksheet.cell(row=rowhistid, column=1, value=idtext)

                    # Increment row number for next write
                    rowhistid += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")
            except NoSuchElementException:
                # Find the next empty cell in the first column
                while worksheet.cell(row=rowhistid, column=1).value:
                    rowhistid += 1
                    
                # Write the text to the cell
                worksheet.cell(row=rowhistid, column=1, value='N/A')

                # Increment row number for next write
                rowhistid += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")                

            try:
                # Find the pre element with class "campo-texto-quebra" inside the div
                label = hist.find_element(By.CSS_SELECTOR, "i.fas.fa-sync-alt")
                labeltext = label.text.strip()

                # Write the text to the worksheet
                if labeltext:
                    # Find the next empty cell in the first column
                    while worksheet.cell(row=rowhistlabel, column=2).value:
                        rowhistlabel += 1
                    
                    # Write the text to the cell
                    worksheet.cell(row=rowhistlabel, column=2, value=labeltext)

                    # Increment row number for next write
                    rowhistlabel += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")
            except NoSuchElementException:
                # Find the next empty cell in the first column
                while worksheet.cell(row=rowhistlabel, column=2).value:
                    rowhistlabel += 1
                    
                # Write the text to the cell
                worksheet.cell(row=rowhistlabel, column=2, value='N/A')

                # Increment row number for next write
                rowhistlabel += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")

            try:
                # Find the pre element with class "campo-texto-quebra" inside the div
                pessoa = hist.find_element(By.CSS_SELECTOR, "b")
                pessoatext = pessoa.text.strip()

                # Write the text to the worksheet
                if pessoatext:
                    # Find the next empty cell in the first column
                    while worksheet.cell(row=rowhistpessoa, column=3).value:
                        rowhistpessoa += 1
                    
                    # Write the text to the cell
                    worksheet.cell(row=rowhistpessoa, column=3, value=pessoatext)

                    # Increment row number for next write
                    rowhistpessoa += 1
                
                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")
            except NoSuchElementException:
                # Find the next empty cell in the first column
                while worksheet.cell(row=rowhistpessoa, column=3).value:
                    rowhistpessoa += 1
                    
                # Write the text to the cell
                worksheet.cell(row=rowhistpessoa, column=3, value='N/A')

                # Increment row number for next write
                rowhistpessoa += 1
                
                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")                

            try:
                # Find the pre element with class "campo-texto-quebra" inside the div
                data = hist.find_element(By.CSS_SELECTOR, "i.fas.fa-sync-alt")
                datatext = data.text.strip()

                # Write the text to the worksheet
                if datatext:
                    # Find the next empty cell in the first column
                    while worksheet.cell(row=rowhistdata, column=4).value:
                        rowhistdata += 1
                    
                    # Write the text to the cell
                    worksheet.cell(row=rowhistdata, column=4, value=datatext)

                    # Increment row number for next write
                    rowhistdata += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")
            except NoSuchElementException:
                while worksheet.cell(row=rowhistdata, column=4).value:
                    rowhistdata += 1
                    
                # Write the text to the cell
                worksheet.cell(row=rowhistdata, column=4, value='N/A')

                # Increment row number for next write
                rowhistdata += 1

                # Save the workbook
                workbook.save("monitoramentoCompleto.xlsx")

        # Wait for the page to finish loading
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
        WebDriverWait(driver, 30).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')

        # Close the current tab and switch back to the main tab
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
    
    # Find the "next" button
    next_button = driver.find_element(By.CSS_SELECTOR, "li.paginate_button.next")
    
    # Check if the button is disabled
    is_disabled = "disabled" in next_button.get_attribute("class")
    if is_disabled:
        # If the button is disabled, break the loop
        break
    
    # Click the "next" button
    next_button.click()

    # Wait for the "carregando" ID to disappear
    wait.until(EC.invisibility_of_element_located((By.ID, 'carregando')))

    # Wait for the page to finish loading
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
    WebDriverWait(driver, 30).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')